import random as rd
import numpy as np
import openpyxl

def F1(x):
    k, l = x
    s1 = 0
    global X
    global Y
    for i in range(5):
        s1 += (k * X[i] + l - Y[i]) ** 2
    return s1

def F2(x):
    m, n = x
    s2=0
    global X
    global Y
    for i in range(5):
        s2 += abs(m*X[i]+n-Y[i])
    return s2

def e(j): #выбор координаты
    ar = np.zeros(2)
    ar[j - 1] = 1
    return ar

def golden_section(f, a, b, eps): #метод золотого сечения
    while (b - a) / 2 > eps:
        c = a + (b - a) * (3 - 5 ** (1 / 2)) / 2
        d = a + (b - a) * (5 ** (1 / 2) - 1) / 2
        if f(c) < f(d):
            b = d
        else:
            a = c
    return a

def coord_boost(f, a, b, eps): #метод покоординатного спуска
    x0 = np.array((rd.uniform(-10, 10), rd.uniform(-10, 10)))
    xk = np.array((rd.uniform(-10, 10), rd.uniform(-10, 10)))
    k = 0
    while np.linalg.norm(x0 - xk) > eps:
        for _ in range(2):
            pk = e(k - int(k / 2) * 2 + 1)
            k += 1
            alpha = golden_section(lambda x: f(xk + x * pk), a, b, eps)
            x0 = xk
            xk = xk + alpha * pk
            sheet.cell(row=1+k, column=1).value = k
            sheet.cell(row=1+k, column=2).value = str(xk[0])+" "+str(xk[1])
            sheet.cell(row=1+k, column=3).value = np.linalg.norm(x0 - xk)
            sheet.cell(row=1+k, column=4).value = f(xk)
            sheet.cell(row=1+k, column=5).value = abs(f(xk)-f(x0))
    return (xk)


def derivative_x(eps, arg, f):
    der_x = (f((10**(-6) + eps, arg)) - f((eps, arg))) / 10**(-6)
    return der_x

def derivative_y(eps, arg, f):
    der_y = (f((arg, eps + 10**(-6))) - f((arg, eps))) / 10**(-6)
    return der_y

def count_alpha(mode, x, gradient, alpha, eps, f):
    if mode == 1: #градиентный метод с постоянным шагом
        return alpha
    if mode == 2: # градиентный метод с дроблением шага
        while f(x - alpha * gradient) > f(x) - 0.5 * alpha * (np.linalg.norm(gradient)) ** 2:
            alpha *= 0.5
        return alpha
    if mode == 3: # метод наискорейшего градиентного спуска
        alpha = golden_section(f=lambda l: f(x - l * gradient), a=-30, b=30, eps=eps)
        return alpha

def Grad(f, eps, mode): #градиентные методы
    alpha = 0.05  # < 2/L, где L = 17.792
    X_prev = np.array([rd.uniform(-10, 10), rd.uniform(-10, 10)])
    gradient = np.array((derivative_x(X_prev[0], X_prev[1], f), derivative_y(X_prev[1], X_prev[0], f)))
    X = X_prev - alpha * gradient
    k = 0
    while np.linalg.norm(X - X_prev) > eps:
        k += 1
        X_prev = X.copy()
        sheet.cell(row=k+1, column=1).value = k
        sheet.cell(row=k+1, column=2).value = str(X_prev[0]) + " " + str(X_prev[1])
        sheet.cell(row=k+1, column=3).value = f(X_prev)
        sheet.cell(row=k+1, column=4).value = np.linalg.norm(gradient)
        gradient = np.array((derivative_x(X_prev[0], X_prev[1], f), derivative_y(X_prev[1], X_prev[0], f)))
        X = X_prev - count_alpha(mode, X_prev, gradient, alpha, eps, f) * gradient
    return X

X = [1, 2, 3, 4, 5]
Y = [-1, -2, -4, 0, -5]
wb = openpyxl.load_workbook(filename='coord_boost_F1.xlsx')
sheet = wb['Sheet1']
sheet.cell(row=1, column=1).value = "n"
sheet.cell(row=1, column=2).value = "xn"
sheet.cell(row=1, column=3).value = "||xn - xn-1||"
sheet.cell(row=1, column=4).value = "f(xn)"
sheet.cell(row=1, column=5).value = "|f(xn) - f(xn-1)|"
minF1 = coord_boost(F1, -10, 10, 10 ** (-6))
wb.save('coord_boost_F1.xlsx')
wb.close()
wb = openpyxl.load_workbook(filename='coord_boost_F2.xlsx')
sheet = wb['Sheet1']
sheet.cell(row=1, column=1).value = "n"
sheet.cell(row=1, column=2).value = "xn"
sheet.cell(row=1, column=3).value = "||xn - xn-1||"
sheet.cell(row=1, column=4).value = "f(xn)"
sheet.cell(row=1, column=5).value = "|f(xn) - f(xn-1)|"
minF2 = coord_boost(F2, -10, 10, 10 ** (-6))
wb.save('coord_boost_F2.xlsx')
wb.close()

wb = openpyxl.load_workbook(filename='grad1.xlsx')
sheet = wb['Sheet1']
sheet.cell(row=1, column=1).value = "n"
sheet.cell(row=1, column=2).value = "xn"
sheet.cell(row=1, column=3).value = "f(xn)"
sheet.cell(row=1, column=4).value = "||f'(xn)||"
min_x, min_y = Grad(F1, 10 ** (-6), 1)
wb.save('grad1.xlsx')
wb.close()

wb = openpyxl.load_workbook(filename='grad2.xlsx')
sheet = wb['Sheet1']
sheet.cell(row=1, column=1).value = "n"
sheet.cell(row=1, column=2).value = "xn"
sheet.cell(row=1, column=3).value = "f(xn)"
sheet.cell(row=1, column=4).value = "||f'(xn)||"
min_x, min_y = Grad(F1, 10 ** (-6), 2)
wb.save('grad2.xlsx')
wb.close()

wb = openpyxl.load_workbook(filename='grad3.xlsx')
sheet = wb['Sheet1']
sheet.cell(row=1, column=1).value = "n"
sheet.cell(row=1, column=2).value = "xn"
sheet.cell(row=1, column=3).value = "f(xn)"
sheet.cell(row=1, column=4).value = "||f'(xn)||"
min_x, min_y = Grad(F1, 10 ** (-6), 3)
wb.save('grad3.xlsx')
wb.close()



